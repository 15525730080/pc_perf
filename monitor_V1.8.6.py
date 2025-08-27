import psutil
import time
import os
from typing import List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import sys
import locale

# 强制使用UTF-8编码
if sys.platform == "win32":
    os.environ["PYTHONUTF8"] = "1"
    locale.setlocale(locale.LC_ALL, "en_US.UTF-8")  # 或 "chs" 中文系统

# 常量定义
# 定义报告文件存储的目录，所有生成的 Excel 报告将保存到该目录下
REPORT_DIR = "Mreport"
# 定义字节到兆字节的转换系数，用于将磁盘 I/O 数据从字节转换为兆字节
MB_CONVERSION = 1024 * 1024
# 定义监控进程性能指标的默认时间间隔，单位为秒，默认值为 0.1 秒
DEFAULT_INTERVAL = 0.1
# 定义监控进程性能指标的默认总时长，单位为秒，默认值为 5 秒
DEFAULT_DURATION = 5
# 定义图表大小的缩放系数，用于调整生成的 Excel 图表的尺寸
CHART_SIZE_MULTIPLIER = 1.7

# GPU支持检测
# 初始化 GPU 支持标志为 False，表示默认情况下不支持 GPU 监控
GPU_SUPPORTED = False
# 初始化 GPU 设备句柄为 None，表示尚未获取到有效的 GPU 设备句柄
GPU_HANDLE = None
try:
    import pynvml
    GPU_SUPPORTED = True
    try:
        pynvml.nvmlInit()
        GPU_HANDLE = pynvml.nvmlDeviceGetHandleByIndex(0)
    except (pynvml.NVMLError_LibraryNotFound, pynvml.NVMLError):
        print("GPU监控不可用，将仅监控CPU和内存")
        GPU_SUPPORTED = False
except ImportError:
    pass


def initialize_monitoring(process_name: str) -> None:
    """初始化进程监控"""
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == process_name:
            proc.cpu_percent(interval=None)
            break


def collect_process_data(proc: psutil.Process, interval: float, 
                        prev_disk_read: int, prev_disk_write: int,
                        prev_net_sent: int, prev_net_recv: int) -> Tuple:
    """收集单个进程的性能数据"""
    raw_cpu = proc.cpu_percent(interval=interval) / psutil.cpu_count(logical=True)
    raw_memory = proc.memory_percent()
    
    # 获取内存使用的具体数值（MB）
    memory_mb = proc.memory_info().rss / (1024 * 1024)
    
    try:
        disk_counters = proc.io_counters()
        disk_read = disk_counters.read_bytes
        disk_write = disk_counters.write_bytes
        
        # 计算磁盘I/O速率，增加最小检测阈值
        disk_read_rate = max(0, (disk_read - prev_disk_read) / interval / MB_CONVERSION) 
        disk_write_rate = max(0, (disk_write - prev_disk_write) / interval / MB_CONVERSION)
        
        # 忽略小于1KB的读写操作
        if disk_read_rate < 0.001:  # 约1KB/s
            disk_read_rate = 0
        if disk_write_rate < 0.001:
            disk_write_rate = 0
            
    except (psutil.AccessDenied, AttributeError):
        disk_read_rate = disk_write_rate = 0
        disk_read = disk_write = 0
    
    # 收集网络IO数据
    try:
        # 获取进程的网络连接
        connections = proc.connections()
        # 汇总系统网络IO数据
        net_io = psutil.net_io_counters()
        net_sent = net_io.bytes_sent
        net_recv = net_io.bytes_recv
        
        # 计算网络IO速率
        net_sent_rate = max(0, (net_sent - prev_net_sent) / interval / MB_CONVERSION)
        net_recv_rate = max(0, (net_recv - prev_net_recv) / interval / MB_CONVERSION)
        
        # 忽略小于1KB的网络传输
        if net_sent_rate < 0.001:
            net_sent_rate = 0
        if net_recv_rate < 0.001:
            net_recv_rate = 0
    except (psutil.AccessDenied, AttributeError):
        net_sent_rate = net_recv_rate = 0
        net_sent = net_recv = 0
    
    return (
        round(raw_cpu, 2),
        round(raw_memory, 2),
        round(memory_mb, 2),  # 返回内存使用的MB值
        round(disk_read_rate, 2),
        round(disk_write_rate, 2),
        disk_read,
        disk_write,
        round(net_sent_rate, 2),
        round(net_recv_rate, 2),
        net_sent,
        net_recv
    )


def update_statistics(stats: dict, cpu: float, memory: float, memory_mb: float,
                     read_rate: float, write_rate: float, 
                     net_sent_rate: float, net_recv_rate: float,
                     gpu: Optional[float] = None) -> None:
    """更新统计数据"""
    stats['max_cpu'] = max(stats['max_cpu'], cpu)
    stats['total_cpu'] += cpu
    stats['cpu_count'] += 1
    
    stats['max_memory'] = max(stats['max_memory'], memory)
    stats['total_memory'] += memory
    stats['memory_count'] += 1
    
    # 更新内存使用MB统计
    stats['max_memory_mb'] = max(stats['max_memory_mb'], memory_mb)
    stats['total_memory_mb'] += memory_mb
    
    stats['max_read'] = max(stats['max_read'], read_rate)
    stats['max_write'] = max(stats['max_write'], write_rate)
    
    # 更新网络IO统计
    stats['max_net_sent'] = max(stats['max_net_sent'], net_sent_rate)
    stats['total_net_sent'] += net_sent_rate
    stats['max_net_recv'] = max(stats['max_net_recv'], net_recv_rate)
    stats['total_net_recv'] += net_recv_rate
    stats['net_count'] += 1
    
    if gpu is not None and GPU_SUPPORTED:
        stats['max_gpu'] = max(stats['max_gpu'], gpu)
        stats['total_gpu'] += gpu
        stats['gpu_count'] += 1


def create_excel_report(process_name: str, data_rows: List[list], stats: dict) -> str:
    """创建Excel报告并返回文件路径"""
    wb = Workbook()
    ws = wb.active
    
    # 设置表头
    headers = ['时间', '进程名称', 'CPU 使用率(%)', '内存使用率(%)', 
              '内存使用量(MB)', '磁盘读取速率(MB/s)', '磁盘写入速率(MB/s)',
              '网络发送速率(MB/s)', '网络接收速率(MB/s)']
    if GPU_SUPPORTED:
        headers.append('GPU 使用率(%)')
    ws.append(headers)
    
    # 写入数据行
    for row in data_rows:
        ws.append(row)
    
    # 添加附加说明，帮助用户理解图表中的多数据系列
    explanation_row = len(data_rows) + 5
    ws.cell(row=explanation_row, column=1, value="图表说明：")
    ws.cell(row=explanation_row + 1, column=1, value="* CPU和内存使用率以百分比(%)显示")
    ws.cell(row=explanation_row + 2, column=1, value="* 内存使用量以MB显示")
    ws.cell(row=explanation_row + 3, column=1, value="* 网络发送和接收速率以MB/s显示")
    
    # 添加统计信息
    stats_row_pos = add_statistics_to_sheet(ws, stats)
    
    # 添加图表区域，使用更适合的位置
    chart_positions = add_charts_to_sheet(ws, len(data_rows), stats_row_pos, GPU_SUPPORTED)
    
    # 调整列宽
    adjust_column_widths(ws)
    
    # 保存文件
    os.makedirs(REPORT_DIR, exist_ok=True)
    file_name = os.path.join(
        REPORT_DIR,
        f"{process_name}_" + time.strftime("%H时%M分%S秒_基础核心性能数据监控报告.xlsx", time.localtime())
    )
    wb.save(file_name)
    return file_name


def add_statistics_to_sheet(ws, stats: dict) -> int:
    """向工作表添加统计信息，返回统计信息末尾行号"""
    stats_col = 11  # 统计信息从K列开始，避免与数据和图表冲突
    stats_row = 2
    ws.cell(row=stats_row, column=stats_col, value='统计信息')
    
    metrics = [
        ('平均 CPU 使用率(%)', (stats['total_cpu'] / stats['cpu_count'])/100 if stats['cpu_count'] > 0 else 0),
        ('峰值 CPU 使用率(%)', stats['max_cpu']/100),
        ('平均内存使用率(%)', (stats['total_memory'] / stats['memory_count'])/100 if stats['memory_count'] > 0 else 0),
        ('峰值内存使用率(%)', stats['max_memory']/100),
        ('平均内存使用量(MB)', round(stats['total_memory_mb'] / stats['memory_count'], 2) if stats['memory_count'] > 0 else 0),
        ('峰值内存使用量(MB)', round(stats['max_memory_mb'], 2)),
        ('最大磁盘读取速率(MB/s)', round(stats['max_read'], 2)),
        ('最大磁盘写入速率(MB/s)', round(stats['max_write'], 2)),
        ('平均网络发送速率(MB/s)', round(stats['total_net_sent'] / stats['net_count'], 2) if stats['net_count'] > 0 else 0),
        ('峰值网络发送速率(MB/s)', round(stats['max_net_sent'], 2)),
        ('平均网络接收速率(MB/s)', round(stats['total_net_recv'] / stats['net_count'], 2) if stats['net_count'] > 0 else 0),
        ('峰值网络接收速率(MB/s)', round(stats['max_net_recv'], 2)),
        ('监控间隔时间(秒)', stats.get('interval', 0.1)),
        ('监控持续时间(秒)', stats.get('duration', 5)),
        ('实际运行时间(秒)', round(stats.get('end_time', 0) - stats.get('start_time', 0), 2))
    ]
    
    for i, (label, value) in enumerate(metrics, start=1):
        ws.cell(row=stats_row + i, column=stats_col, value=label)
        cell = ws.cell(row=stats_row + i, column=stats_col + 1, value=value)
        if '使用率' in label:
            cell.number_format = '0.00%'  # 使用百分比格式会自动显示正确值
        else:
            cell.number_format = '0.00'
    
    last_row = stats_row + len(metrics)
    
    if GPU_SUPPORTED and stats['gpu_count'] > 0:
        last_row += 1
        ws.cell(row=last_row, column=stats_col, value='平均 GPU 使用率(%)').number_format = '0.00%'
        ws.cell(row=last_row, column=stats_col + 1, value=(stats['total_gpu'] / stats['gpu_count'])/100)
        last_row += 1
        ws.cell(row=last_row, column=stats_col, value='峰值 GPU 使用率(%)').number_format = '0.00%'
        ws.cell(row=last_row, column=stats_col + 1, value=stats['max_gpu']/100)
    
    return last_row


def add_charts_to_sheet(ws, data_rows_count: int, stats_last_row: int, gpu_supported: bool) -> dict:
    """添加所有图表到工作表，并返回图表位置信息"""
    
    # 计算各图表的位置
    positions = {}
    
    # 获取图表说明的位置
    explanation_row = data_rows_count + 5
    # 将图表放在图表说明下方第3行
    chart_start_row = explanation_row + 5  # 说明占3行 + 空2行
    
    # 创建合并的综合性能图表
    combined_chart = create_combined_performance_chart(ws, data_rows_count, gpu_supported)
    positions['combined_chart'] = {'row': chart_start_row, 'col': 'A'}
    ws.add_chart(combined_chart, f'A{chart_start_row}')
    
    # 冻结窗格以便于查看
    ws.freeze_panes = ws['A2']
    
    return positions


def create_combined_performance_chart(ws, data_rows_count: int, gpu_supported: bool) -> LineChart:
    """创建综合性能监控图表，包含CPU、内存和网络IO数据"""
    chart = LineChart()
    chart.title = '综合性能监控'
    chart.y_axis.title = '使用率/速率/内存量'
    chart.x_axis.title = '时间'
    
    # 定义要包含的列
    # CPU和内存使用率在3-4列，内存MB在5列，网络IO在8-9列
    included_cols = [3, 4, 5, 8, 9]  # CPU, 内存百分比, 内存MB, 网络发送, 网络接收
    
    if gpu_supported:
        included_cols.append(10)  # GPU
    
    # 为每组数据创建引用
    data_refs = []
    for col in included_cols:
        data_refs.append(Reference(ws, min_col=col, min_row=1, max_col=col, max_row=data_rows_count + 1))
    
    # 类别（时间轴）
    cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count + 1)
    
    # 添加所有数据系列到图表
    for data_ref in data_refs:
        chart.add_data(data_ref, titles_from_data=True)
    
    chart.set_categories(cats)
    
    # 设置图表样式
    chart.style = 13
    chart.legend.position = 't'
    chart.y_axis.majorGridlines = None
    chart.x_axis.tickLblSkip = min(5, max(2, data_rows_count // 10))
    chart.width = 24 * CHART_SIZE_MULTIPLIER
    chart.height = 12 * CHART_SIZE_MULTIPLIER
    
    # 样式细节
    chart.x_axis.number_format = 'hh:mm:ss'
    chart.x_axis.majorTickMark = 'none'
    
    # 添加辅助线来区分CPU/内存和网络IO的量级差异
    if data_rows_count > 0:
        try:
            # 获取前几行数据用于估计网络IO值范围
            sample_size = min(data_rows_count, 10)  # 最多取10行样本
            max_net_values = []
            
            # 遍历表格中的前几行获取网络IO数据，用于估计线条粗细
            for row in range(2, 2 + sample_size):
                net_send_value = ws.cell(row=row, column=8).value or 0
                net_recv_value = ws.cell(row=row, column=9).value or 0
                max_net_values.append(max(net_send_value, net_recv_value))
                
            # 计算平均值作为参考
            avg_net_value = sum(max_net_values) / len(max_net_values) if max_net_values else 0
            
            # 如果网络IO值较大，调整网络IO线条样式
            if avg_net_value > 10:  # 大于10MB/s时调整
                for i in range(3, 5):  # 调整网络发送和接收线条
                    if i < len(chart.series):
                        chart.series[i].graphicalProperties.line.width = 30000  # 加粗线条
        except Exception:
            # 忽略任何样式调整错误，确保图表能够正常生成
            pass
    
    return chart


def adjust_column_widths(ws) -> None:
    """调整工作表列宽"""
    for column in ws.columns:
        max_length = max(
            len(str(cell.value)) for cell in column
        )
        adjusted_width = (max_length * 1.2) + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def monitor_process(process_name: str, interval: float = DEFAULT_INTERVAL, duration: float = DEFAULT_DURATION) -> None:
    """
    监控指定进程的性能指标
    
    :param process_name: 要监控的进程名称(如'notepad.exe')
    :param interval: 监控间隔时间(秒)，默认0.1秒
    :param duration: 监控总时长(秒)，默认5秒
    """
    # 检查目标进程是否存在
    start_time = time.time()
    timeout = 5  # 5秒超时
    
    while True:
        target_pids = [p.pid for p in psutil.process_iter(['name']) if p.info['name'] == process_name]
        if target_pids:
            break
            
        if time.time() - start_time > timeout:
            raise Exception(f"错误: 目标进程 {process_name} 未运行，请先启动该进程")
            
        time.sleep(0.1)
        
    initialize_monitoring(process_name)
    
    end_time = time.time() + duration
    data_rows = []
    
    # 初始化时获取初始IO值
    initial_disk_read = initial_disk_write = 0
    initial_net_sent = initial_net_recv = 0
    
    # 初始化网络IO计数器
    net_io = psutil.net_io_counters()
    initial_net_sent = net_io.bytes_sent
    initial_net_recv = net_io.bytes_recv
    
    target_pids = [p.pid for p in psutil.process_iter(['name']) if p.info['name'] == process_name]
    for pid in target_pids:
        try:
            proc = psutil.Process(pid)
            io = proc.io_counters()
            initial_disk_read = io.read_bytes
            initial_disk_write = io.write_bytes
            break
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    
    prev_disk_read = initial_disk_read
    prev_disk_write = initial_disk_write
    prev_net_sent = initial_net_sent
    prev_net_recv = initial_net_recv
    
    stats = {
        'max_cpu': 0, 'total_cpu': 0, 'cpu_count': 0,
        'max_memory': 0, 'total_memory': 0, 'memory_count': 0,
        'max_memory_mb': 0, 'total_memory_mb': 0,  # 新增内存MB统计
        'max_read': 0, 'max_write': 0,
        'max_net_sent': 0, 'total_net_sent': 0,
        'max_net_recv': 0, 'total_net_recv': 0, 'net_count': 0,
        'max_gpu': 0, 'total_gpu': 0, 'gpu_count': 0,
        'interval': interval, 'duration': duration,
        'start_time': time.time()
    }
    
    # 获取目标进程的pid列表，避免每次循环都遍历所有进程
    target_pids = [p.pid for p in psutil.process_iter(['name']) if p.info['name'] == process_name]
    
    try:
        while time.time() < end_time:
            # 实时进程存活检测（PID存在性检查+进程名匹配）
            target_pids = []
            try:
                target_pids = [p.pid for p in psutil.process_iter(['name', 'pid']) 
                              if p.info['name'] == process_name and psutil.pid_exists(p.pid)]
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

            if not target_pids:
                print(f"[WARNING] 进程 {process_name} 已终止，提前结束监控（剩余时间：{round(end_time - time.time(), 2)}秒）")
                stats['end_time'] = time.time()
                break

            loop_start = time.time()  # 记录循环开始时间
            
            for pid in target_pids:
                try:
                    proc = psutil.Process(pid)
                    proc.cpu_percent(interval=None)
                    cpu, memory, memory_mb, read_rate, write_rate, prev_disk_read, prev_disk_write, \
                    net_sent_rate, net_recv_rate, prev_net_sent, prev_net_recv = collect_process_data(
                        proc, interval, prev_disk_read, prev_disk_write, prev_net_sent, prev_net_recv
                    )
                    
                    current_time = time.strftime("%H:%M:%S", time.localtime())
                    row_data = [current_time, process_name, cpu, memory, memory_mb, read_rate, write_rate, net_sent_rate, net_recv_rate]
                    
                    if GPU_SUPPORTED:
                        gpu = round(pynvml.nvmlDeviceGetUtilizationRates(GPU_HANDLE).gpu, 2)
                        row_data.append(gpu)
                        update_statistics(stats, cpu, memory, memory_mb, read_rate, write_rate, net_sent_rate, net_recv_rate, gpu)
                    else:
                        update_statistics(stats, cpu, memory, memory_mb, read_rate, write_rate, net_sent_rate, net_recv_rate)
                    
                    data_rows.append(row_data)
                    print_monitoring_data(process_name, cpu, memory, memory_mb, read_rate, write_rate, net_sent_rate, net_recv_rate)
                    
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
                    print(f"[ERROR] 进程数据采集失败: {str(e)}")
                    stats['end_time'] = time.time()
                    break
            
            # 精确控制间隔时间
            elapsed = time.time() - loop_start
            if elapsed < interval:
                time.sleep(interval - elapsed)
        
        # 确保结束时间被记录
        if 'end_time' not in stats:
            stats['end_time'] = time.time()
            
    except Exception as e:
        # 捕获所有异常，确保能生成报告
        print(f"[ERROR] 监控过程发生错误: {str(e)}")
        if 'end_time' not in stats:
            stats['end_time'] = time.time()
    
    finally:
        # 即使发生错误也尝试生成报告
        try:
            if data_rows:  # 只有在有数据时才生成报告
                file_path = create_excel_report(process_name, data_rows, stats)
                print(f"已生成报告: {file_path}")
                print_report_summary(stats)
            else:
                print("没有收集到有效数据，无法生成报告")
        except Exception as report_error:
            print(f"[ERROR] 生成报告时发生错误: {str(report_error)}")
        
        # 关闭GPU资源
        if GPU_SUPPORTED:
            try:
                pynvml.nvmlShutdown()
            except:
                pass


def print_monitoring_data(process_name: str, cpu: float, memory: float, memory_mb: float, 
                         read_rate: float, write_rate: float,
                         net_sent_rate: float, net_recv_rate: float) -> None:
    """打印监控数据到控制台"""
    print(f"进程名称: {process_name}")
    print(f"CPU 使用率: {cpu}%")
    print(f"内存使用率: {memory}%")
    print(f"内存使用量: {memory_mb} MB")
    print(f"磁盘读取速率: {read_rate} MB/s")
    print(f"磁盘写入速率: {write_rate} MB/s")
    print(f"网络发送速率: {net_sent_rate} MB/s")
    print(f"网络接收速率: {net_recv_rate} MB/s")
    if GPU_SUPPORTED:
        try:
            gpu = round(pynvml.nvmlDeviceGetUtilizationRates(GPU_HANDLE).gpu, 2)
            print(f"GPU 使用率: {gpu}%")
        except:
            pass
    print("-" * 30)


def print_report_summary(stats: dict) -> None:
    """打印报告摘要到控制台"""
    print(f"监控期间峰值 CPU 使用率: {round(stats['max_cpu'], 2)}%")
    print(f"监控期间平均 CPU 使用率: {round(stats['total_cpu'] / stats['cpu_count'], 2) if stats['cpu_count'] > 0 else 0}%")
    print(f"监控期间峰值内存使用率: {round(stats['max_memory'], 2)}%")
    print(f"监控期间平均内存使用率: {round(stats['total_memory'] / stats['memory_count'], 2) if stats['memory_count'] > 0 else 0}%")
    print(f"监控期间峰值内存使用量: {round(stats['max_memory_mb'], 2)} MB")
    print(f"监控期间平均内存使用量: {round(stats['total_memory_mb'] / stats['memory_count'], 2) if stats['memory_count'] > 0 else 0} MB")
    print(f"监控期间最大磁盘读取速率: {round(stats['max_read'], 2)} MB/s")
    print(f"监控期间最大磁盘写入速率: {round(stats['max_write'], 2)} MB/s")
    print(f"监控期间最大网络发送速率: {round(stats['max_net_sent'], 2)} MB/s")
    print(f"监控期间平均网络发送速率: {round(stats['total_net_sent'] / stats['net_count'], 2) if stats['net_count'] > 0 else 0} MB/s")
    print(f"监控期间最大网络接收速率: {round(stats['max_net_recv'], 2)} MB/s")
    print(f"监控期间平均网络接收速率: {round(stats['total_net_recv'] / stats['net_count'], 2) if stats['net_count'] > 0 else 0} MB/s")
    if GPU_SUPPORTED and stats['gpu_count'] > 0:
        print(f"监控期间峰值 GPU 使用率: {round(stats['max_gpu'], 2)}%")
        print(f"监控期间平均 GPU 使用率: {round(stats['total_gpu'] / stats['gpu_count'], 2)}%")


if __name__ == "__main__":
    """ 常用进程名称
    1. TTVoice.exe; 极速版-TT语音
    2. HeyboxChat.exe; 黑盒语音
    3. oopz.exe; Oopz
    4. TT语音.exe; 营收PC-TT
    5. VALORANT-Win64-Shipping.exe; 瓦洛兰特
    """
    monitor_process('Cursor.exe', interval=1, duration=300)
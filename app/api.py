# coding: utf-8
"""
FastAPI 路由层 — 精简、轻量。

所有响应统一格式：{"code": 200, "msg": <data>}
"""
import base64
import os
import platform
import shutil
import time
import traceback
from datetime import datetime
from pathlib import Path

import asyncio
from fastapi import FastAPI
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from pydantic import BaseModel
from starlette.requests import Request
from starlette.responses import JSONResponse, RedirectResponse, FileResponse

from app.db import TaskCollection, ComparisonReportCollection, LabelCollection, create_tables
from app.log import log as logger
from app.task_handle import TaskHandle
from app.util import DataCollect
from app.core.device_manager import (
    DeviceManager,
    get_platform_capabilities,
    DEVICE_TYPE_PC,
    DEVICE_TYPE_ANDROID,
    DEVICE_TYPE_IOS,
    DEVICE_TYPE_HARMONY,
)

# ── 应用初始化 ────────────────────────────────────────────────
app = FastAPI(title="client-perf", version="2.0.0")

BASE_DIR = Path(__file__).parent.parent / "test_result"
BASE_DIR.mkdir(exist_ok=True)
# 同时兼容 /test_result 和旧的 /static 访问路径。
app.mount("/test_result", StaticFiles(directory=str(BASE_DIR)), name="test_result")
app.mount("/static", StaticFiles(directory=str(BASE_DIR)), name="static")


# ── 统一响应 ──────────────────────────────────────────────────

def ok(data=None):
    return JSONResponse({"code": 200, "msg": data})


def err(msg: str, code: int = 500):
    return JSONResponse({"code": code, "msg": msg})


# ── 全局异常中间件 ────────────────────────────────────────────

@app.middleware("http")
async def _error_handler(request: Request, call_next):
    try:
        return await call_next(request)
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))


# ── 报告导出功能 ──────────────────────────────────────────────

# 报告保存目录
REPORT_DIR = Path(__file__).parent.parent / "test_result" / "excel_reports"
REPORT_DIR.mkdir(exist_ok=True, parents=True)

COMPARISON_REPORT_DIR = Path(__file__).parent.parent / "test_result" / "comparison_reports"
COMPARISON_REPORT_DIR.mkdir(exist_ok=True, parents=True)


def create_excel_report(task_name: str, data: list[dict], save_dir: str) -> str:
    """
    导出单个任务的 Excel 报告
    """
    try:
        # 创建 Workbook
        wb = Workbook()
        
        # 为每个指标创建一个工作表
        for item in data:
            sheet_name = item.get("name", "未知")
            ws = wb.create_sheet(title=sheet_name)
            
            # 写入表头和数据
            values = item.get("value", [])
            if values:
                # 收集所有可能的字段
                all_fields = set()
                for row in values:
                    all_fields.update(row.keys())
                headers = sorted(list(all_fields))
                
                # 写入表头
                ws.append(headers)
                
                # 写入数据
                for row in values:
                    row_data = [row.get(h) for h in headers]
                    ws.append(row_data)
            else:
                # 无数据时写入提示
                ws.append(["无数据"])
        
        # 删除默认的工作表
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # 生成文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{task_name}_{timestamp}.xlsx"
        file_path = REPORT_DIR / filename
        
        # 保存 Excel
        wb.save(file_path)
        
        logger.info(f"Excel 报告导出成功: {file_path}")
        return str(file_path)
    except Exception as e:
        logger.error(f"导出 Excel 报告失败: {e}")
        raise


async def create_comparison_excel(data: dict, report_name: str) -> str:
    """
    导出任务对比报告为 Excel
    """
    try:
        # 创建 Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "对比结果"
        
        # 写入表头
        headers = [
            "任务名称", "版本", "基准", "CPU 均值(%)", "内存均值(MB)", 
            "FPS 均值", "GPU 均值(%)", "线程数均值", "句柄数均值", 
            "磁盘读取均值(MB/s)", "磁盘写入均值(MB/s)", "网络发送均值(MB/s)", "网络接收均值(MB/s)"
        ]
        ws.append(headers)
        
        # 写入数据
        for task in data.get("tasks", []):
            row = [
                task.get("name", ""),
                task.get("version", ""),
                "是" if task.get("id") == data.get("base_task", {}).get("id") else "否",
                task.get("avg", {}).get("cpu_avg", "—"),
                task.get("avg", {}).get("memory_avg", "—"),
                task.get("avg", {}).get("fps_avg", "—"),
                task.get("avg", {}).get("gpu_avg", "—"),
                task.get("avg", {}).get("threads_avg", "—"),
                task.get("avg", {}).get("handles_avg", "—"),
                task.get("avg", {}).get("disk_read_avg", "—"),
                task.get("avg", {}).get("disk_write_avg", "—"),
                task.get("avg", {}).get("net_sent_avg", "—"),
                task.get("avg", {}).get("net_recv_avg", "—"),
            ]
            ws.append(row)
        
        # 为每个任务添加原始数据工作表
        for task in data.get("tasks", []):
            task_name = task.get("name", "未知任务")
            task_id = task.get("id", "")
            sheet_name = f"{task_name}_{task_id}"
            
            # 限制工作表名称长度
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            # 创建新工作表
            task_ws = wb.create_sheet(title=sheet_name)
            
            # 写入原始数据
            task_data = task.get("data", {})
            metrics = [
                ("cpu", "CPU 使用率 (%)"),
                ("memory", "内存使用量 (MB)"),
                ("fps", "FPS"),
                ("gpu", "GPU 使用率 (%)"),
                ("threads", "线程数"),
                ("handles", "句柄数"),
                ("disk_read", "磁盘读取 (MB/s)"),
                ("disk_write", "磁盘写入 (MB/s)"),
                ("net_sent", "网络发送 (MB/s)"),
                ("net_recv", "网络接收 (MB/s)"),
            ]
            
            row_idx = 1
            for metric_key, metric_name in metrics:
                metric_data = task_data.get(metric_key, [])
                if metric_data:
                    # 写入指标名称
                    task_ws.cell(row=row_idx, column=1, value=metric_name)
                    row_idx += 1
                    
                    # 写入表头
                    task_ws.cell(row=row_idx, column=1, value="时间")
                    task_ws.cell(row=row_idx, column=2, value="值")
                    row_idx += 1
                    
                    # 写入数据
                    for item in metric_data:
                        time_val = item.get("time")
                        value_val = item.get("value")
                        if time_val:
                            # 转换时间戳为可读时间
                            time_str = datetime.fromtimestamp(time_val).strftime("%Y-%m-%d %H:%M:%S")
                            task_ws.cell(row=row_idx, column=1, value=time_str)
                            task_ws.cell(row=row_idx, column=2, value=value_val)
                            row_idx += 1
                    
                    # 空行分隔不同指标
                    row_idx += 1
        
        # 生成文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{report_name}_{timestamp}.xlsx"
        file_path = COMPARISON_REPORT_DIR / filename
        
        # 保存 Excel
        wb.save(file_path)
        
        logger.info(f"对比报告导出成功: {file_path}")
        return str(file_path)
    except Exception as e:
        logger.error(f"导出对比报告失败: {e}")
        raise


async def export_label_comparison_excel_func(data: dict, report_name: str) -> str:
    """
    导出标签对比报告为 Excel
    """
    try:
        # 创建 Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "标签对比结果"
        
        # 写入表头
        headers = [
            "任务名称", "版本", "标签名称", "开始时间", "结束时间", 
            "CPU 均值(%)", "内存均值(MB)", "FPS 均值", "GPU 均值(%)", 
            "线程数均值", "句柄数均值", "磁盘读取均值(MB/s)", 
            "磁盘写入均值(MB/s)", "网络发送均值(MB/s)", "网络接收均值(MB/s)"
        ]
        ws.append(headers)
        
        # 写入数据
        for task in data.get("tasks", []):
            row = [
                task.get("name", ""),
                task.get("version", ""),
                task.get("label_name", ""),
                datetime.fromtimestamp(task.get("start_ts", 0)).strftime("%Y-%m-%d %H:%M:%S"),
                datetime.fromtimestamp(task.get("end_ts", 0)).strftime("%Y-%m-%d %H:%M:%S"),
                task.get("avg", {}).get("cpu_avg", "—"),
                task.get("avg", {}).get("memory_avg", "—"),
                task.get("avg", {}).get("fps_avg", "—"),
                task.get("avg", {}).get("gpu_avg", "—"),
                task.get("avg", {}).get("threads_avg", "—"),
                task.get("avg", {}).get("handles_avg", "—"),
                task.get("avg", {}).get("disk_read_avg", "—"),
                task.get("avg", {}).get("disk_write_avg", "—"),
                task.get("avg", {}).get("net_sent_avg", "—"),
                task.get("avg", {}).get("net_recv_avg", "—"),
            ]
            ws.append(row)
        
        # 为每个标签添加原始数据工作表
        for task in data.get("tasks", []):
            task_name = task.get("name", "未知任务")
            label_name = task.get("label_name", "未知标签")
            sheet_name = f"{task_name}_{label_name}"
            
            # 限制工作表名称长度
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            # 创建新工作表
            task_ws = wb.create_sheet(title=sheet_name)
            
            # 写入原始数据
            task_data = task.get("data", {})
            metrics = [
                ("cpu", "CPU 使用率 (%)"),
                ("memory", "内存使用量 (MB)"),
                ("fps", "FPS"),
                ("gpu", "GPU 使用率 (%)"),
                ("threads", "线程数"),
                ("handles", "句柄数"),
                ("disk_read", "磁盘读取 (MB/s)"),
                ("disk_write", "磁盘写入 (MB/s)"),
                ("net_sent", "网络发送 (MB/s)"),
                ("net_recv", "网络接收 (MB/s)"),
            ]
            
            row_idx = 1
            for metric_key, metric_name in metrics:
                # 标签数据的结构可能与任务数据不同，需要特殊处理
                if metric_key in task_data:
                    metric_data = task_data[metric_key]
                    timestamps = task_data.get("timestamps", [])
                    
                    if metric_data and timestamps:
                        # 写入指标名称
                        task_ws.cell(row=row_idx, column=1, value=metric_name)
                        row_idx += 1
                        
                        # 写入表头
                        task_ws.cell(row=row_idx, column=1, value="时间")
                        task_ws.cell(row=row_idx, column=2, value="值")
                        row_idx += 1
                        
                        # 写入数据
                        for i, value in enumerate(metric_data):
                            if i < len(timestamps):
                                time_val = timestamps[i]
                                # 转换时间戳为可读时间
                                time_str = datetime.fromtimestamp(time_val).strftime("%Y-%m-%d %H:%M:%S")
                                task_ws.cell(row=row_idx, column=1, value=time_str)
                                task_ws.cell(row=row_idx, column=2, value=value)
                                row_idx += 1
                        
                        # 空行分隔不同指标
                        row_idx += 1
        
        # 生成文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"{report_name}_{timestamp}.xlsx"
        file_path = COMPARISON_REPORT_DIR / filename
        
        # 保存 Excel
        wb.save(file_path)
        
        logger.info(f"标签对比报告导出成功: {file_path}")
        return str(file_path)
    except Exception as e:
        logger.error(f"导出标签对比报告失败: {e}")
        raise


# ── 生命周期 ──────────────────────────────────────────────────

@app.on_event("startup")
async def _startup():
    await create_tables()
    # 定期检查僵尸 monitor 进程
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    scheduler = AsyncIOScheduler()
    scheduler.add_job(_cleanup_zombie_monitors, "interval", seconds=60)
    scheduler.start()
    app.state.scheduler = scheduler


@app.on_event("shutdown")
async def _shutdown():
    if hasattr(app.state, "scheduler"):
        app.state.scheduler.shutdown(wait=False)
    try:
        from app.core.ios_tools import TunnelManager
        TunnelManager.stop_all_tunnels()
    except ImportError:
        pass


async def _cleanup_zombie_monitors():
    """定期清理已停止任务的残留 monitor 进程"""
    import psutil
    pids_alive = {p.pid for p in psutil.process_iter()}
    monitor_pids = await TaskCollection.get_all_stop_task_monitor_pid()
    for mpid in monitor_pids:
        if mpid and mpid in pids_alive:
            logger.info(f"[cleanup] kill zombie monitor_pid={mpid}")
            TaskHandle.stop_handle(mpid)


# ── 路由 ──────────────────────────────────────────────────────

@app.get("/")
def index():
    return RedirectResponse(url="/test_result/index.html")


# ── 设备 ──────────────────────────────────────────────────────

@app.get("/get_devices/")
async def get_devices():
    try:
        return ok(DeviceManager.get_all_devices())
    except Exception as e:
        return err(str(e))


@app.get("/platform_capabilities/")
async def platform_capabilities():
    try:
        return ok(get_platform_capabilities())
    except Exception as e:
        return err(str(e))


@app.get("/system_info/")
async def system_info(device_type: str = "pc", device_id: str = None):
    try:
        info = await DeviceManager.get_device_sys_info(device_type, device_id or "")
        return ok(info)
    except Exception as e:
        return err(str(e))


@app.get("/get_pids/")
async def get_pids(
    is_print_tree: bool = False,
    device_type: str = "pc",
    device_id: str = None,
):
    try:
        if device_type == DEVICE_TYPE_ANDROID:
            from app.core.android_tools import android_packages
            return ok(await android_packages(device_id))
        elif device_type == DEVICE_TYPE_IOS:
            from app.core.ios_tools import ios_apps
            return ok(await ios_apps(device_id))
        elif device_type == DEVICE_TYPE_HARMONY:
            from app.core.harmony_tools import harmony_packages
            return ok(await harmony_packages(device_id))
        else:
            from app.core.pc_tools import process_tree, pids
            return ok(await process_tree() if is_print_tree else await pids())
    except Exception as e:
        return err(str(e))


@app.get("/get_device_apps/")
async def get_device_apps(device_type: str, device_id: str):
    try:
        apps = await DeviceManager.get_device_apps_async(device_type, device_id)
        return ok(apps)
    except Exception as e:
        return err(str(e))


@app.get("/pid_img/")
async def pid_img(pid: int = 0, device_type: str = "pc", device_id: str = None):
    try:
        img = await DeviceManager.take_screenshot(device_type, device_id or "", pid)
        if img:
            return base64.b64encode(img).decode()
        return ""
    except Exception as e:
        return err(str(e))


# ── 任务管理 ──────────────────────────────────────────────────

@app.get("/get_all_task/")
async def get_all_task():
    return ok(await TaskCollection.get_all_task())


@app.get("/run_task/")
async def run_task(
    pid: int = 0,
    pid_name: str = "",
    task_name: str = "",
    include_child: bool = False,
    device_type: str = "pc",
    device_id: str = None,
    package_name: str = None,
):
    try:
        task_id, file_dir = await TaskCollection.create_task(
            pid, pid_name, str(BASE_DIR), task_name, include_child,
            device_type=device_type, device_id=device_id, package_name=package_name,
        )
        handle = TaskHandle(
            serialno=device_id or platform.node(),
            file_dir=file_dir,
            task_id=task_id,
            platform_name=platform.system() if device_type == "pc" else device_type,
            target_pid=pid,
            include_child=include_child,
            device_type=device_type,
            device_id=device_id,
            package_name=package_name,
        )
        handle.start()
        return ok()
    except Exception as e:
        return err(str(e))


@app.get("/stop_task/")
async def stop_task(task_id: int):
    try:
        task = await TaskCollection.stop_task(task_id)
        TaskHandle.stop_handle(task.get("monitor_pid"))
        return ok()
    except Exception as e:
        return err(str(e))


@app.get("/task_status/")
async def task_status(task_id: int):
    try:
        task = await TaskCollection.get_item_task(task_id)
        return ok(task.get("status"))
    except Exception as e:
        return err(str(e))


@app.get("/result/")
async def task_result(task_id: int):
    try:
        task = await TaskCollection.get_item_task(task_id)
        data = await DataCollect(task["file_dir"]).get_all_data()
        return ok(data)
    except Exception as e:
        return err(str(e))


@app.get("/delete_task/")
async def delete_task(task_id: int):
    try:
        item = await TaskCollection.delete_task(task_id)
        if item.get("file_dir") and os.path.exists(item["file_dir"]):
            shutil.rmtree(item["file_dir"], ignore_errors=True)
        return ok()
    except Exception as e:
        return err(str(e))


@app.get("/change_task_name/")
async def change_task_name(task_id: int, new_name: str):
    try:
        task = await TaskCollection.change_task_name(task_id, new_name)
        return ok(f"已重命名为: {task['name']}")
    except Exception as e:
        return err(str(e))


@app.get("/set_task_version/")
async def set_task_version(task_id: int, version: str):
    try:
        await TaskCollection.set_task_version(task_id, version)
        return ok(f"已设置版本: {version}")
    except Exception as e:
        return err(str(e))


@app.get("/set_task_baseline/")
async def set_task_baseline(task_id: int, is_baseline: bool = True):
    try:
        await TaskCollection.set_task_baseline(task_id, is_baseline)
        return ok("已设置基线" if is_baseline else "已取消基线")
    except Exception as e:
        return err(str(e))


@app.get("/get_baseline_task/")
async def get_baseline_task():
    try:
        task = await TaskCollection.get_baseline_task()
        return ok(task)
    except Exception as e:
        return err(str(e))


# ── Excel 导出 ────────────────────────────────────────────────

@app.get("/export_excel/")
async def export_excel(task_id: int):
    try:
        task = await TaskCollection.get_item_task(task_id)
        data = await DataCollect(task["file_dir"]).get_all_data(is_format=False)
        file_path = create_excel_report(
            task.get("name") or f"任务{task_id}", data, task["file_dir"]
        )
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))


# ── 对比分析 ──────────────────────────────────────────────────

@app.get("/compare_tasks/")
async def compare_tasks(task_ids: str, base_task_id: int = None):
    try:
        from app.comparison import TaskComparison
        id_list = [int(x.strip()) for x in task_ids.split(",") if x.strip()]
        data = await TaskComparison.create_comparison(id_list, base_task_id)
        return ok(data)
    except Exception as e:
        return err(str(e))


@app.get("/export_comparison_excel/")
async def export_comparison_excel(
    task_ids: str,
    base_task_id: int = None,
    report_name: str = None,
):
    try:
        from app.comparison import TaskComparison

        id_list = [int(x.strip()) for x in task_ids.split(",") if x.strip()]
        data = await TaskComparison.create_comparison(id_list, base_task_id)
        name = report_name or f"性能对比_{time.strftime('%Y%m%d_%H%M%S')}"
        file_path = await create_comparison_excel(data, name)

        report = await ComparisonReportCollection.create_report(
            name=name,
            task_ids=id_list,
            base_task_id=base_task_id or id_list[0],
            description=f"对比任务: {task_ids}",
        )
        await ComparisonReportCollection.update_report(
            report["id"], report_path=file_path
        )
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))


@app.get("/get_comparison_reports/")
async def get_comparison_reports():
    try:
        return ok(await ComparisonReportCollection.get_all_reports())
    except Exception as e:
        return err(str(e))


@app.get("/delete_comparison_report/")
async def delete_comparison_report(report_id: int):
    try:
        report = await ComparisonReportCollection.delete_report(report_id)
        rp = report.get("report_path")
        if rp and os.path.exists(rp):
            os.remove(rp)
        return ok("已删除")
    except Exception as e:
        return err(str(e))


@app.get("/export_label_comparison_excel/")
async def export_label_comparison_excel(label_ids: str, report_name: str = None):
    """
    导出标签对比报告为 Excel
    """
    try:
        ids = [int(x.strip()) for x in label_ids.split(",") if x.strip()]
        if len(ids) < 2:
            return err("至少需要两个标签", 400)
        
        labels = [await LabelCollection.get_label(lid) for lid in ids]
        
        async def _load_sliced(label: dict):
            info = await TaskCollection.get_item_task(label["task_id"])
            raw  = await DataCollect(info["file_dir"]).get_all_data(is_format=False)
            sliced = [
                {
                    "name": item["name"],
                    "value": [
                        v for v in item.get("value", [])
                        if label["start_ts"] <= v.get("time", 0) <= label["end_ts"]
                    ],
                }
                for item in raw
            ]
            return info, sliced, label
        
        results = await asyncio.gather(*[_load_sliced(lb) for lb in labels])
        
        # 指标映射：metric_key -> (csv_stem, csv_column_prefix)
        metric_map = {
            "cpu":        ("cpu",          "cpu_usage"),
            "memory":     ("memory",       "process_memory_usage"),
            "fps":        ("fps",          "fps"),
            "gpu":        ("gpu",          "gpu"),
            "threads":    ("process_info", "num_threads"),
            "handles":    ("process_info", "num_handles"),
            "disk_read":  ("disk_io",      "disk_read_rate"),
            "disk_write": ("disk_io",      "disk_write_rate"),
            "net_sent":   ("network_io",   "net_sent_rate"),
            "net_recv":   ("network_io",   "net_recv_rate"),
        }
        
        def _extract_values(data: list[dict], stem: str, col_prefix: str) -> list[float]:
            """从 get_all_data() 结果中提取指定指标的数值列表。"""
            for item in data:
                if item.get("name") != stem:
                    continue
                vals: list[float] = []
                for row in item.get("value", []):
                    for k, v in row.items():
                        if k.startswith(col_prefix) and isinstance(v, (int, float)):
                            vals.append(float(v))
                            break
                return vals
            return []
        
        def _calc_avg(vals: list[float]) -> float | None:
            return round(sum(vals) / len(vals), 4) if vals else None
        
        def _build_task_avg(data: list[dict]) -> dict[str, float | None]:
            """返回 {metric_avg_key: value} 形如 {"cpu_avg": 12.3, ...}"""
            result: dict[str, float | None] = {}
            for metric, (stem, col) in metric_map.items():
                vals = _extract_values(data, stem, col)
                result[f"{metric}_avg"] = _calc_avg(vals)
            return result
        
        tasks_result = []
        for info, sliced_data, label in results:
            tasks_result.append({
                "id":          info["id"],
                "name":        info.get("name", ""),
                "version":     info.get("version", ""),
                "label_id":    label["id"],
                "label_name":  label["name"],
                "label_color": label["color"],
                "start_ts":    label["start_ts"],
                "end_ts":      label["end_ts"],
                "avg":         _build_task_avg(sliced_data),
            })
        
        base = tasks_result[0]
        for t in tasks_result:
            diff, pct = {}, {}
            for k, v in t["avg"].items():
                bv = base["avg"].get(k)
                if v is not None and bv is not None:
                    d = round(v - bv, 4)
                    diff[k] = d
                    pct[k]  = round(d / bv * 100, 2) if bv != 0 else None
                else:
                    diff[k] = pct[k] = None
            t["diff"] = diff
            t["pct"]  = pct
        
        data = {
            "base_task": {
                "id":         base["id"],
                "name":       base["name"],
                "label_name": base["label_name"],
            },
            "tasks": tasks_result,
        }
        
        name = report_name or f"标签对比_{time.strftime('%Y%m%d_%H%M%S')}"
        file_path = await export_label_comparison_excel_func(data, name)
        
        report = await ComparisonReportCollection.create_report(
            name=name,
            task_ids=[t["id"] for t in tasks_result],
            base_task_id=base["id"],
            description=f"对比标签: {label_ids}",
        )
        await ComparisonReportCollection.update_report(
            report["id"], report_path=file_path
        )
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))


# ── 高级对比分析 ──────────────────────────────────────────────

@app.get("/advanced_compare_tasks/")
async def advanced_compare_tasks(task_ids: str, base_task_id: int = None):
    """统计显著性 + 异常值检测 + 瓶颈分析"""
    import math
    try:
        id_list = [int(x.strip()) for x in task_ids.split(",") if x.strip()]
        if len(id_list) < 2:
            return err("高级分析需要至少两个任务", 400)

        base_id = base_task_id or id_list[0]
        cmp_id  = id_list[1] if id_list[0] == base_id else id_list[0]

        base_info = await TaskCollection.get_item_task(base_id)
        cmp_info  = await TaskCollection.get_item_task(cmp_id)
        base_data = await DataCollect(base_info["file_dir"]).get_all_data()
        cmp_data  = await DataCollect(cmp_info["file_dir"]).get_all_data()

        metric_map = {
            "cpu":        ("cpu",          "cpu_usage"),
            "memory":     ("memory",       "process_memory_usage"),
            "fps":        ("fps",          "fps"),
            "gpu":        ("gpu",          "gpu"),
            "threads":    ("process_info", "num_threads"),
            "disk_read":  ("disk_io",      "disk_read_rate"),
            "disk_write": ("disk_io",      "disk_write_rate"),
            "net_sent":   ("network_io",   "net_sent_rate"),
            "net_recv":   ("network_io",   "net_recv_rate"),
        }
        better_smaller = {"cpu", "memory", "disk_read", "disk_write", "net_sent", "net_recv"}

        def extract(data, dk, vk):
            vals = []
            for item in data:
                if item.get("name") == dk:
                    for v in item.get("value", []):
                        raw = v.get(vk)
                        if raw not in (None, "", "-"):
                            try:
                                vals.append(float(raw))
                            except (ValueError, TypeError):
                                pass
            return vals

        def stats(arr):
            n = len(arr)
            if n == 0:
                return 0.0, 0.0, 0.0
            mean = sum(arr) / n
            var  = sum((x - mean) ** 2 for x in arr) / max(n - 1, 1)
            return mean, var, n

        significance = {}
        outliers_result = {}
        bottlenecks = []

        for metric, (dk, vk) in metric_map.items():
            bv = extract(base_data, dk, vk)
            cv = extract(cmp_data,  dk, vk)
            if not bv or not cv:
                continue

            bm, bvar, bn = stats(bv)
            cm, cvar, cn = stats(cv)
            diff = cm - bm
            pct  = (diff / bm * 100) if bm != 0 else 0

            se = ((bvar / bn) + (cvar / cn)) ** 0.5 if (bvar / bn + cvar / cn) > 0 else 1e-9
            t  = abs(diff / se) if se > 0 else 0
            p  = 2 * (1 - min(0.9999, 0.5 * (1 + math.erf(t / math.sqrt(2)))))
            significant = p < 0.05

            significance[metric] = {
                "base_mean": round(bm, 4),
                "compare_mean": round(cm, 4),
                "diff": round(diff, 4),
                "percent_change": round(pct, 2),
                "p_value": round(p, 4),
                "confidence": round((1 - p) * 100, 2),
                "is_significant": significant,
            }

            # 异常值（Z-score > 2.5）
            combined = bv + cv
            mc = sum(combined) / len(combined)
            sc = (sum((x - mc) ** 2 for x in combined) / max(len(combined) - 1, 1)) ** 0.5
            outlier_list = [
                {"value": round(v, 4), "z_score": round(abs((v - mc) / sc) if sc else 0, 2), "is_high": v > mc}
                for v in cv if sc and abs((v - mc) / sc) > 2.5
            ]
            if outlier_list:
                outliers_result[metric] = {
                    "outliers": outlier_list[:20],
                    "summary": f"检测到 {len(outlier_list)} 个异常值",
                }

            # 瓶颈
            if significant and abs(pct) > 20:
                is_worse = (pct > 0) if metric in better_smaller else (pct < 0)
                bottlenecks.append({
                    "metric": metric,
                    "percent_change": round(pct, 2),
                    "is_worse": is_worse,
                })

        bottlenecks.sort(key=lambda x: abs(x["percent_change"]), reverse=True)

        return ok({
            "base_task":    {"id": base_id, "name": base_info.get("name", "")},
            "compare_task": {"id": cmp_id,  "name": cmp_info.get("name", "")},
            "advanced_analysis": {
                "statistical_significance": significance,
                "outliers": outliers_result,
                "bottleneck_analysis": {
                    "potential_bottlenecks": bottlenecks,
                    "summary": f"共发现 {len(bottlenecks)} 个潜在性能瓶颈",
                },
            },
        })
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))


# ── 标签管理（区间标签）────────────────────────────────────────

@app.get("/get_labels/")
async def get_labels(task_id: int | None = None):
    """获取区间标签；传 task_id 则只返回该任务的标签，否则返回全部。"""
    try:
        if task_id is not None:
            return ok(await LabelCollection.get_labels_by_task(task_id))
        return ok(await LabelCollection.get_all_labels())
    except Exception as e:
        return err(str(e))


class LabelCreateBody(BaseModel):
    task_id: int
    name: str
    start_ts: float
    end_ts: float
    color: str = "#3b6ef0"
    note: str = ""


@app.post("/create_label/")
async def create_label(body: LabelCreateBody):
    """在任务时间轴上创建一个区间标签（接收 JSON body）"""
    try:
        label = await LabelCollection.create_label(
            task_id=body.task_id, name=body.name,
            start_ts=body.start_ts, end_ts=body.end_ts,
            color=body.color, note=body.note,
        )
        return ok(label)
    except Exception as e:
        return err(str(e))


@app.post("/update_label/")
async def update_label(
    label_id: int,
    name: str = None,
    start_ts: float = None,
    end_ts: float = None,
    color: str = None,
    note: str = None,
):
    """更新区间标签"""
    try:
        label = await LabelCollection.update_label(
            label_id=label_id, name=name,
            start_ts=start_ts, end_ts=end_ts,
            color=color, note=note,
        )
        return ok(label)
    except Exception as e:
        return err(str(e))


@app.delete("/delete_label/")
async def delete_label(label_id: int):
    """删除区间标签"""
    try:
        return ok(await LabelCollection.delete_label(label_id))
    except Exception as e:
        return err(str(e))


# ── 区间对比 ──────────────────────────────────────────────────

@app.get("/compare_labels/")
async def compare_labels(label_ids: str):
    """
    基于区间标签对比。
    label_ids: 逗号分隔的 label id，每个 label 对应一个任务的一段区间。
    """
    try:
        from app.comparison import _build_task_avg
        import asyncio

        ids = [int(x.strip()) for x in label_ids.split(",") if x.strip()]
        if len(ids) < 2:
            return err("至少需要两个标签", 400)

        labels = [await LabelCollection.get_label(lid) for lid in ids]

        async def _load_sliced(label: dict):
            info = await TaskCollection.get_item_task(label["task_id"])
            raw  = await DataCollect(info["file_dir"]).get_all_data(is_format=False)
            sliced = [
                {
                    "name": item["name"],
                    "value": [
                        v for v in item.get("value", [])
                        if label["start_ts"] <= v.get("time", 0) <= label["end_ts"]
                    ],
                }
                for item in raw
            ]
            return info, sliced, label

        results = await asyncio.gather(*[_load_sliced(lb) for lb in labels])

        # 指标映射：metric_key -> (csv_stem, csv_column_prefix)
        metric_map = {
            "cpu":        ("cpu",          "cpu_usage"),
            "memory":     ("memory",       "process_memory_usage"),
            "fps":        ("fps",          "fps"),
            "gpu":        ("gpu",          "gpu"),
            "threads":    ("process_info", "num_threads"),
            "handles":    ("process_info", "num_handles"),
            "disk_read":  ("disk_io",      "disk_read_rate"),
            "disk_write": ("disk_io",      "disk_write_rate"),
            "net_sent":   ("network_io",   "net_sent_rate"),
            "net_recv":   ("network_io",   "net_recv_rate"),
        }

        def extract_detailed_data(sliced_data):
            """提取标签区间内的详细数据"""
            data = {"timestamps": []}
            # 初始化各指标的数据数组
            for metric in metric_map:
                data[metric] = []
            
            # 先获取时间戳
            for item in sliced_data:
                if item.get("name") == "cpu":  # 任意选择一个指标获取时间戳
                    for v in item.get("value", []):
                        data["timestamps"].append(v.get("time", 0))
                    break
            
            # 提取各指标的数据
            for metric, (stem, col_prefix) in metric_map.items():
                for item in sliced_data:
                    if item.get("name") != stem:
                        continue
                    for v in item.get("value", []):
                        for k, val in v.items():
                            if k.startswith(col_prefix) and isinstance(val, (int, float)):
                                data[metric].append(float(val))
                                break
                        else:
                            data[metric].append(None)
                    break
            
            return data

        tasks_result = []
        for info, sliced_data, label in results:
            tasks_result.append({
                "id":          info["id"],
                "name":        info.get("name", ""),
                "version":     info.get("version", ""),
                "label_id":    label["id"],
                "label_name":  label["name"],
                "label_color": label["color"],
                "start_ts":    label["start_ts"],
                "end_ts":      label["end_ts"],
                "avg":         _build_task_avg(sliced_data),
                "data":        extract_detailed_data(sliced_data),
            })

        base = tasks_result[0]
        for t in tasks_result:
            diff, pct = {}, {}
            for k, v in t["avg"].items():
                bv = base["avg"].get(k)
                if v is not None and bv is not None:
                    d = round(v - bv, 4)
                    diff[k] = d
                    pct[k]  = round(d / bv * 100, 2) if bv != 0 else None
                else:
                    diff[k] = pct[k] = None
            t["diff"] = diff
            t["pct"]  = pct

        return ok({
            "base_task": {
                "id":         base["id"],
                "name":       base["name"],
                "label_name": base["label_name"],
            },
            "tasks": tasks_result,
        })
    except Exception as e:
        logger.error(traceback.format_exc())
        return err(str(e))

"""
性能数据对比报表生成模块

用于生成性能对比的Excel报表
"""
import os
import time
from typing import List, Dict, Any, Optional
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabel
from openpyxl.utils import get_column_letter
from app.log import log as logger
from app.database import TaskCollection
from app.util import DataCollect

# 常量定义
CHART_WIDTH = 30  # 图表宽度
CHART_HEIGHT = 15  # 图表高度
BETTER_IF_SMALLER = {"cpu", "memory", "disk_write", "disk_read", "net_sent", "net_recv"}  # 值越小越好的指标

# 样式定义
HEADER_FONT = Font(name='Arial', size=12, bold=True)
SUBHEADER_FONT = Font(name='Arial', size=11, bold=True)
NORMAL_FONT = Font(name='Arial', size=10)
BETTER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色，表示更好
WORSE_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # 红色，表示更差
NEUTRAL_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # 黄色，表示变化不大
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


async def create_comparison_excel(tasks_data: Dict[str, Any], report_name: Optional[str] = None, 
                           save_dir: Optional[str] = None) -> str:
    """
    创建性能对比Excel报表
    
    Args:
        tasks_data: 任务对比数据
        report_name: 报告名称，如果不提供则自动生成
        save_dir: 保存目录，如果不提供则保存在默认位置
        
    Returns:
        Excel文件路径
    """
    logger.info("开始创建对比报表Excel")
    
    wb = Workbook()
    
    # 创建总览工作表
    ws_summary = wb.active
    ws_summary.title = "对比总览"
    
    # 创建详细数据工作表
    ws_details = wb.create_sheet(title="详细数据")
    
    # 创建趋势图工作表
    ws_charts = wb.create_sheet(title="趋势对比")
    
    # 填充总览工作表
    populate_summary_sheet(ws_summary, tasks_data)
    
    # 填充详细数据工作表
    populate_details_sheet(ws_details, tasks_data)
    
    # 填充趋势图工作表
    populate_charts_sheet(ws_charts, tasks_data)
    
    # 设置列宽
    sheet_list = [ws_summary, ws_details, ws_charts]
    for ws in sheet_list:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 生成文件名和保存路径
    if not report_name:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        report_name = f"性能对比报告_{timestamp}"
    
    if not save_dir:
        save_dir = os.path.join("test_result", "comparison_reports")
        
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, f"{report_name}.xlsx")
    
    try:
        wb.save(file_path)
        logger.info(f"对比报表保存成功：{file_path}")
    except Exception as e:
        logger.error(f"保存对比报表失败：{str(e)}")
        raise
        
    return file_path


def populate_summary_sheet(ws, tasks_data: Dict[str, Any]) -> None:
    """填充总览工作表"""
    # 添加标题
    ws.merge_cells('A1:G1')
    ws['A1'] = "性能对比总览"
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加基本信息
    row = 3
    ws.cell(row=row, column=1, value="对比基准任务：")
    if "base_task" in tasks_data and tasks_data["base_task"]:
        base_task = tasks_data["base_task"]
        ws.cell(row=row, column=2, value=f"{base_task.get('name')} (ID: {base_task.get('id')})")
        if "version" in base_task:
            ws.cell(row=row, column=2, value=f"{base_task.get('name')} - {base_task.get('version')} (ID: {base_task.get('id')})")
    row += 1
    
    ws.cell(row=row, column=1, value="对比任务数：")
    ws.cell(row=row, column=2, value=len(tasks_data.get("tasks", [])))
    row += 1
    
    ws.cell(row=row, column=1, value="生成时间：")
    ws.cell(row=row, column=2, value=time.strftime("%Y-%m-%d %H:%M:%S"))
    row += 2
    
    # 添加任务列表
    ws.cell(row=row, column=1, value="任务列表").font = HEADER_FONT
    row += 1
    
    headers = ["序号", "任务ID", "任务名称", "版本", "开始时间", "结束时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1
    
    # 添加任务信息
    for i, task in enumerate(tasks_data.get("tasks", []), 1):
        ws.cell(row=row, column=1, value=i).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=2, value=task.get("id")).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=3, value=task.get("name"))
        ws.cell(row=row, column=4, value=task.get("version"))
        
        start_time = task.get("start_time")
        if start_time:
            if isinstance(start_time, str):
                ws.cell(row=row, column=5, value=start_time)
            else:
                ws.cell(row=row, column=5, value=start_time.strftime("%Y-%m-%d %H:%M:%S") if hasattr(start_time, "strftime") else str(start_time))
        
        end_time = task.get("end_time")
        if end_time:
            if isinstance(end_time, str):
                ws.cell(row=row, column=6, value=end_time)
            else:
                ws.cell(row=row, column=6, value=end_time.strftime("%Y-%m-%d %H:%M:%S") if hasattr(end_time, "strftime") else str(end_time))
        
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        # 标记基准任务
        if tasks_data.get("base_task") and task.get("id") == tasks_data["base_task"].get("id"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        
        row += 1
    
    row += 2
    
    # 添加指标对比摘要
    ws.cell(row=row, column=1, value="性能指标对比摘要").font = HEADER_FONT
    row += 1
    
    summary_headers = ["指标", "平均差异", "平均变化率(%)", "变化趋势"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1
    
    # 指标名称映射
    metric_name_map = {
        "cpu": "CPU使用率",
        "memory": "内存使用量",
        "fps": "FPS帧率",
        "gpu": "GPU使用率",
        "threads": "线程数",
        "handles": "句柄数",
        "disk_read": "磁盘读取",
        "disk_write": "磁盘写入",
        "net_sent": "网络发送",
        "net_recv": "网络接收"
    }
    
    # 添加指标对比摘要
    summary = tasks_data.get("summary", {})
    for metric, data in summary.items():
        ws.cell(row=row, column=1, value=metric_name_map.get(metric, metric))
        ws.cell(row=row, column=2, value=data.get("avg_diff")).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=3, value=data.get("avg_percent_change")).alignment = CENTER_ALIGNMENT
        
        # 根据指标改善方向判断变化趋势
        is_better = data.get("is_better", False)
        if is_better:
            trend_text = "改善"
            ws.cell(row=row, column=4, value=trend_text).fill = BETTER_FILL
        else:
            trend_text = "下降"
            ws.cell(row=row, column=4, value=trend_text).fill = WORSE_FILL
        ws.cell(row=row, column=4).alignment = CENTER_ALIGNMENT
        
        for col in range(1, len(summary_headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        row += 1
    
    row += 2
    
    # 添加说明
    ws.cell(row=row, column=1, value="说明：").font = SUBHEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="1. 对于CPU使用率、内存使用量、磁盘I/O和网络I/O，值越小越好")
    row += 1
    ws.cell(row=row, column=1, value="2. 对于FPS帧率，值越大越好")
    row += 1
    ws.cell(row=row, column=1, value="3. 平均差异 = 对比任务 - 基准任务")
    row += 1
    ws.cell(row=row, column=1, value="4. 平均变化率(%) = (对比任务 - 基准任务) / 基准任务 × 100%")


def populate_details_sheet(ws, tasks_data: Dict[str, Any]) -> None:
    """填充详细数据工作表"""
    # 添加标题
    ws.merge_cells('A1:F1')
    ws['A1'] = "性能指标详细对比"
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 3
    
    # 添加各项指标的详细对比
    metrics = [
        {"key": "cpu", "name": "CPU使用率", "unit": "%"},
        {"key": "memory", "name": "内存使用量", "unit": "MB"},
        {"key": "fps", "name": "FPS帧率", "unit": "帧"},
        {"key": "gpu", "name": "GPU使用率", "unit": "%"},
        {"key": "threads", "name": "线程数", "unit": "个"},
        {"key": "handles", "name": "句柄数", "unit": "个"},
        {"key": "disk_read", "name": "磁盘读取速率", "unit": "MB/s"},
        {"key": "disk_write", "name": "磁盘写入速率", "unit": "MB/s"},
        {"key": "net_sent", "name": "网络发送速率", "unit": "MB/s"},
        {"key": "net_recv", "name": "网络接收速率", "unit": "MB/s"}
    ]
    
    for metric in metrics:
        # 添加指标标题
        ws.cell(row=row, column=1, value=f"{metric['name']}对比").font = HEADER_FONT
        row += 1
        
        # 添加表头
        headers = ["任务ID", "任务名称", "平均值", "峰值"]
        if metric["key"] == "fps":  # FPS特殊处理，显示最低值而不是峰值
            headers[3] = "最低值"
            
        for i, task in enumerate(tasks_data.get("tasks", [])):
            if i > 0:  # 对于非基准任务，添加对比列
                headers.append(f"相对差异")
                headers.append(f"变化率(%)")
                
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
        row += 1
        
        # 添加任务数据
        tasks = tasks_data.get("tasks", [])
        base_task = tasks_data.get("base_task")
        base_task_id = base_task.get("id") if base_task else None
        
        for task in tasks:
            task_id = task.get("id")
            task_avg = task.get("avg", {})
            
            col = 1
            # 任务ID
            ws.cell(row=row, column=col, value=task_id).alignment = CENTER_ALIGNMENT
            col += 1
            
            # 任务名称
            ws.cell(row=row, column=col, value=task.get("name"))
            col += 1
            
            # 平均值
            avg_key = f"{metric['key']}_avg"
            ws.cell(row=row, column=col, value=task_avg.get(avg_key)).alignment = CENTER_ALIGNMENT
            col += 1
            
            # 峰值或最低值
            if metric["key"] == "fps":
                peak_key = f"{metric['key']}_min"
                peak_label = "最低值"
            else:
                peak_key = f"{metric['key']}_max"
                peak_label = "峰值"
            ws.cell(row=row, column=col, value=task_avg.get(peak_key)).alignment = CENTER_ALIGNMENT
            col += 1
            
            # 如果不是基准任务，添加对比数据
            if task_id != base_task_id:
                # 在metrics中查找该指标的diff和percent_change
                metric_data = tasks_data.get("metrics", {}).get(metric["key"], {})
                diff_list = metric_data.get("diff", [])
                percent_list = metric_data.get("percent_change", [])
                
                # 找到当前任务的diff和percent_change
                diff = next((item["value"] for item in diff_list if item["task_id"] == task_id), None)
                percent = next((item["value"] for item in percent_list if item["task_id"] == task_id), None)
                
                # 相对差异
                cell = ws.cell(row=row, column=col, value=diff)
                cell.alignment = CENTER_ALIGNMENT
                # 根据值的大小和指标类型设置颜色
                if diff is not None:
                    is_better_if_smaller = metric["key"] in BETTER_IF_SMALLER
                    if (is_better_if_smaller and diff < 0) or (not is_better_if_smaller and diff > 0):
                        cell.fill = BETTER_FILL
                    elif (is_better_if_smaller and diff > 0) or (not is_better_if_smaller and diff < 0):
                        cell.fill = WORSE_FILL
                    else:
                        cell.fill = NEUTRAL_FILL
                col += 1
                
                # 变化率(%)
                cell = ws.cell(row=row, column=col, value=percent)
                cell.alignment = CENTER_ALIGNMENT
                # 根据值的大小和指标类型设置颜色
                if percent is not None:
                    is_better_if_smaller = metric["key"] in BETTER_IF_SMALLER
                    if (is_better_if_smaller and percent < 0) or (not is_better_if_smaller and percent > 0):
                        cell.fill = BETTER_FILL
                    elif (is_better_if_smaller and percent > 0) or (not is_better_if_smaller and percent < 0):
                        cell.fill = WORSE_FILL
                    else:
                        cell.fill = NEUTRAL_FILL
            
            # 为所有单元格添加边框
            for c in range(1, col + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            
            # 标记基准任务
            if task_id == base_task_id:
                for c in range(1, col + 1):
                    ws.cell(row=row, column=c).fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            
            row += 1
        
        # 在每个指标对比后添加空行
        row += 1


def populate_charts_sheet(ws, tasks_data: Dict[str, Any]) -> None:
    """填充趋势图工作表"""
    # 添加标题
    ws.merge_cells('A1:G1')
    ws['A1'] = "性能趋势对比图"
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建条形图，对比各指标的平均值
    row = 3
    ws.cell(row=row, column=1, value="各指标平均值对比").font = HEADER_FONT
    row += 1
    
    # 准备图表数据
    metrics = [
        {"key": "cpu_avg", "name": "CPU平均使用率(%)"},
        {"key": "memory_avg", "name": "内存平均使用(MB)"},
        {"key": "fps_avg", "name": "FPS平均帧率(帧)"},
        {"key": "gpu_avg", "name": "GPU平均使用率(%)"},
        {"key": "threads_avg", "name": "平均线程数(个)"},
        {"key": "handles_avg", "name": "平均句柄数(个)"},
        {"key": "disk_read_avg", "name": "磁盘平均读取(MB/s)"},
        {"key": "disk_write_avg", "name": "磁盘平均写入(MB/s)"},
        {"key": "net_sent_avg", "name": "网络平均发送(MB/s)"},
        {"key": "net_recv_avg", "name": "网络平均接收(MB/s)"}
    ]
    
    # 添加数据表格
    # 表头
    ws.cell(row=row, column=1, value="指标")
    for i, task in enumerate(tasks_data.get("tasks", [])):
        ws.cell(row=row, column=i+2, value=task.get("name"))
    row += 1
    
    # 填充数据行
    chart_data_rows = []
    for metric in metrics:
        # 检查是否有任务包含此指标
        has_metric = False
        for task in tasks_data.get("tasks", []):
            if metric["key"] in task.get("avg", {}):
                has_metric = True
                break
                
        if not has_metric:
            continue
            
        chart_data_rows.append(row)
        ws.cell(row=row, column=1, value=metric["name"])
        
        for i, task in enumerate(tasks_data.get("tasks", [])):
            avg = task.get("avg", {}).get(metric["key"])
            ws.cell(row=row, column=i+2, value=avg)
        row += 1
    
    # 创建条形图
    for i, metric_row in enumerate(chart_data_rows):
        metric_name = ws.cell(row=metric_row, column=1).value
        chart = BarChart()
        chart.title = f"{metric_name}对比"
        chart.style = 10
        chart.x_axis.title = "任务"
        chart.y_axis.title = metric_name.split('(')[1].replace(')', '') if '(' in metric_name else ""
        
        # 设置数据范围
        data = Reference(ws, min_row=metric_row, max_row=metric_row, 
                       min_col=2, max_col=len(tasks_data.get("tasks", [])) + 1)
        cats = Reference(ws, min_row=row - len(chart_data_rows), max_row=row - len(chart_data_rows),
                       min_col=2, max_col=len(tasks_data.get("tasks", [])) + 1)
        
        chart.add_data(data)
        chart.set_categories(cats)
        
        # 设置图表大小和位置
        chart.width = CHART_WIDTH
        chart.height = CHART_HEIGHT
        
        # 计算图表位置
        chart_row = row + i * (CHART_HEIGHT // 20 + 1)
        ws.add_chart(chart, f"A{chart_row}")
    
    # 更新行指针
    row = row + len(chart_data_rows) * (CHART_HEIGHT // 20 + 1) 
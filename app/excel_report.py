import os
import time
import logging
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import openpyxl.styles

from app.log import log as logger

# 常量定义
MB_CONVERSION = 1024 * 1024
CHART_SIZE_MULTIPLIER = 1.7

def format_timestamp(timestamp):
    """将时间戳格式化为时间字符串"""
    return time.strftime("%H:%M:%S", time.localtime(int(timestamp)))

def create_excel_report(task_name: str, task_data: List[Dict], file_dir: str) -> str:
    """
    创建Excel报告并返回文件路径
    
    :param task_name: 任务名称
    :param task_data: 性能数据
    :param file_dir: 文件保存目录
    :return: 生成的Excel文件路径
    """
    logger.info(f"开始创建Excel报表，任务名称: {task_name}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "性能数据"
    
    # 冻结首行和首列，便于浏览
    ws.freeze_panes = 'B2'
    
    # 提取所有数据系列
    series_data = {}
    time_data = []
    
    for data_item in task_data:
        name = data_item.get('name')
        values = data_item.get('value', [])
        
        logger.info(f"处理数据类型: {name}, 数据点数量: {len(values)}")
        
        # 收集时间数据
        if not time_data and values:
            time_data = [format_timestamp(v.get('time', 0)) for v in values]
        
        # 根据数据类型收集值
        if name == 'cpu':
            series_data['cpu_usage'] = [v.get('cpu_usage(%)', '-') for v in values]
            series_data['cpu_usage_all'] = [v.get('cpu_usage_all(%)', '-') for v in values]
        elif name == 'memory':
            series_data['memory_usage'] = [v.get('process_memory_usage(M)', '-') for v in values]
        elif name == 'fps':
            series_data['fps'] = [v.get('fps(帧)', '-') for v in values]
        elif name == 'gpu':
            series_data['gpu'] = [v.get('gpu(%)', '-') for v in values]
        elif name == 'process_info':
            series_data['threads'] = [v.get('num_threads(个)', '-') for v in values]
            series_data['handles'] = [v.get('num_handles(个)', '-') for v in values]
        # 添加对新指标的收集
        elif name == 'disk_io':
            series_data['disk_read_rate'] = [v.get('disk_read_rate(MB/s)', '-') for v in values]
            series_data['disk_write_rate'] = [v.get('disk_write_rate(MB/s)', '-') for v in values]
        elif name == 'network_io':
            series_data['net_sent_rate'] = [v.get('net_sent_rate(MB/s)', '-') for v in values]
            series_data['net_recv_rate'] = [v.get('net_recv_rate(MB/s)', '-') for v in values]
    
    # 设置表头
    headers = ['时间', 'CPU使用率(%)', 'CPU总使用率(%)', '内存使用量(MB)', 'FPS(帧)']
    if 'gpu' in series_data:
        headers.append('GPU使用率(%)')
    headers.extend(['线程数', '句柄数'])
    # 添加新的表头
    headers.extend(['磁盘读取速率(MB/s)', '磁盘写入速率(MB/s)', '网络发送速率(MB/s)', '网络接收速率(MB/s)'])
    
    # 记录表头
    logger.info(f"表头: {headers}")
    
    ws.append(headers)
    
    # 写入数据行
    data_rows_count = 0
    for i in range(len(time_data)):
        try:
            row = [time_data[i]]
            row.append(series_data.get('cpu_usage', [])[i] if i < len(series_data.get('cpu_usage', [])) else '')
            row.append(series_data.get('cpu_usage_all', [])[i] if i < len(series_data.get('cpu_usage_all', [])) else '')
            row.append(series_data.get('memory_usage', [])[i] if i < len(series_data.get('memory_usage', [])) else '')
            row.append(series_data.get('fps', [])[i] if i < len(series_data.get('fps', [])) else '')
            
            if 'gpu' in series_data:
                row.append(series_data.get('gpu', [])[i] if i < len(series_data.get('gpu', [])) else '')
            
            row.append(series_data.get('threads', [])[i] if i < len(series_data.get('threads', [])) else '')
            row.append(series_data.get('handles', [])[i] if i < len(series_data.get('handles', [])) else '')
            
            # 添加新指标数据
            row.append(series_data.get('disk_read_rate', [])[i] if i < len(series_data.get('disk_read_rate', [])) else '')
            row.append(series_data.get('disk_write_rate', [])[i] if i < len(series_data.get('disk_write_rate', [])) else '')
            row.append(series_data.get('net_sent_rate', [])[i] if i < len(series_data.get('net_sent_rate', [])) else '')
            row.append(series_data.get('net_recv_rate', [])[i] if i < len(series_data.get('net_recv_rate', [])) else '')
            
            ws.append(row)
            data_rows_count += 1
        except Exception as e:
            logger.error(f"写入数据行 {i} 时出错: {str(e)}")
    
    logger.info(f"数据行写入完成，共 {data_rows_count} 行")
    
    # 创建数据概览工作表
    ws_summary = wb.create_sheet(title="数据概览")
    add_summary_sheet(wb, ws_summary, task_name, series_data, data_rows_count)
    
    # 调整列宽
    adjust_column_widths(ws)
    
    # 保存文件
    report_dir = os.path.join(file_dir, "excel_reports")
    os.makedirs(report_dir, exist_ok=True)
    
    file_name = os.path.join(
        report_dir,
        f"{task_name}_{time.strftime('%H时%M分%S秒_性能数据报告.xlsx', time.localtime())}"
    )
    
    try:
        # 尝试不带图表保存一次，确保基本数据能正常保存
        wb.save(file_name)
        logger.info(f"基本数据保存成功: {file_name}")
        
        # 为主数据表添加统计信息
        stats_row = add_statistics_to_sheet(ws, series_data, data_rows_count)
        logger.info(f"统计信息添加完成")
        
        # 尝试添加所有图表
        try:
            add_charts_to_sheet(ws, data_rows_count + 1, stats_row, 'gpu' in series_data)
            logger.info("所有图表添加成功")
            wb.save(file_name)
        except Exception as chart_error:
            logger.error(f"添加图表时出错: {str(chart_error)}")
            # 已经有基本数据版本了，可以忽略图表错误
            
    except Exception as save_error:
        logger.error(f"保存Excel文件时出错: {str(save_error)}")
        raise
        
    logger.info(f"Excel报表创建完成: {file_name}")
    return file_name

def add_summary_sheet(wb, ws, task_name, series_data, data_rows_count):
    """添加数据概览工作表"""
    ws.cell(row=1, column=1, value=f"任务名称: {task_name}")
    ws.cell(row=2, column=1, value=f"数据采集时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=3, column=1, value=f"数据点数量: {data_rows_count}")
    
    row = 5
    ws.cell(row=row, column=1, value="性能指标统计")
    row += 1
    
    # 计算统计值
    stats = {}
    
    for key, values in series_data.items():
        # 过滤掉非数字值
        numeric_values = []
        for v in values:
            try:
                if v != '-' and v is not None:
                    numeric_values.append(float(v))
            except (ValueError, TypeError):
                pass
                
        if numeric_values:
            stats[f'max_{key}'] = max(numeric_values)
            stats[f'avg_{key}'] = sum(numeric_values) / len(numeric_values)
            stats[f'min_{key}'] = min(numeric_values)
    
    # 添加统计行
    metrics = []
    
    if 'avg_cpu_usage' in stats:
        metrics.extend([
            ('CPU平均使用率(%)', stats.get('avg_cpu_usage', 0)),
            ('CPU峰值使用率(%)', stats.get('max_cpu_usage', 0)),
            ('CPU最低使用率(%)', stats.get('min_cpu_usage', 0)),
        ])
        
    if 'avg_memory_usage' in stats:
        metrics.extend([
            ('平均内存使用量(MB)', stats.get('avg_memory_usage', 0)),
            ('峰值内存使用量(MB)', stats.get('max_memory_usage', 0)),
            ('最低内存使用量(MB)', stats.get('min_memory_usage', 0)),
        ])
        
    if 'avg_fps' in stats:
        metrics.extend([
            ('平均FPS(帧)', stats.get('avg_fps', 0)),
            ('峰值FPS(帧)', stats.get('max_fps', 0)),
            ('最低FPS(帧)', stats.get('min_fps', 0)),
        ])
        
    if 'avg_gpu' in stats:
        metrics.extend([
            ('平均GPU使用率(%)', stats.get('avg_gpu', 0)),
            ('峰值GPU使用率(%)', stats.get('max_gpu', 0)),
            ('最低GPU使用率(%)', stats.get('min_gpu', 0)),
        ])
        
    if 'avg_threads' in stats:
        metrics.extend([
            ('平均线程数', stats.get('avg_threads', 0)),
            ('峰值线程数', stats.get('max_threads', 0)),
        ])
        
    if 'avg_handles' in stats:
        metrics.extend([
            ('平均句柄数', stats.get('avg_handles', 0)),
            ('峰值句柄数', stats.get('max_handles', 0)),
        ])
    
    # 添加新指标的统计信息
    if 'avg_disk_read_rate' in stats:
        metrics.extend([
            ('平均磁盘读取速率(MB/s)', stats.get('avg_disk_read_rate', 0)),
            ('峰值磁盘读取速率(MB/s)', stats.get('max_disk_read_rate', 0)),
        ])
        
    if 'avg_disk_write_rate' in stats:
        metrics.extend([
            ('平均磁盘写入速率(MB/s)', stats.get('avg_disk_write_rate', 0)),
            ('峰值磁盘写入速率(MB/s)', stats.get('max_disk_write_rate', 0)),
        ])
        
    if 'avg_net_sent_rate' in stats:
        metrics.extend([
            ('平均网络发送速率(MB/s)', stats.get('avg_net_sent_rate', 0)),
            ('峰值网络发送速率(MB/s)', stats.get('max_net_sent_rate', 0)),
        ])
        
    if 'avg_net_recv_rate' in stats:
        metrics.extend([
            ('平均网络接收速率(MB/s)', stats.get('avg_net_recv_rate', 0)),
            ('峰值网络接收速率(MB/s)', stats.get('max_net_recv_rate', 0)),
        ])
    
    # 写入统计数据
    for i, (label, value) in enumerate(metrics):
        ws.cell(row=row + i, column=1, value=label)
        cell = ws.cell(row=row + i, column=2, value=round(value, 2))
    
    # 调整列宽
    adjust_column_widths(ws)

def add_statistics_to_sheet(ws, series_data, data_length):
    """向工作表添加统计信息，返回统计信息末尾行号"""
    # 将统计信息列从J列(10)改为O列(15)，确保不会与数据重叠
    stats_col = 15  
    stats_row = 2
    
    # 设置统计信息标题样式
    title_cell = ws.cell(row=stats_row, column=stats_col, value='统计信息')
    title_cell.font = openpyxl.styles.Font(bold=True, size=12)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # 合并标题单元格
    ws.merge_cells(start_row=stats_row, start_column=stats_col, end_row=stats_row, end_column=stats_col + 1)
    
    # 为标题添加背景色
    title_cell.fill = openpyxl.styles.PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # 计算统计值
    stats = {}
    
    for key, values in series_data.items():
        # 过滤掉非数字值
        numeric_values = []
        for v in values:
            try:
                if v != '-' and v is not None:
                    numeric_values.append(float(v))
            except (ValueError, TypeError):
                pass
                
        if numeric_values:
            stats[f'max_{key}'] = max(numeric_values)
            stats[f'avg_{key}'] = sum(numeric_values) / len(numeric_values)
    
    # 添加统计行
    metrics = []
    
    # 添加分类标题：CPU相关
    if 'avg_cpu_usage' in stats:
        metrics.append(("CPU相关指标", None, True))  # 添加标题行
        metrics.extend([
            ('平均CPU使用率(%)', stats.get('avg_cpu_usage', 0), False),
            ('峰值CPU使用率(%)', stats.get('max_cpu_usage', 0), False),
        ])
    
    # 添加分类标题：内存相关
    if 'avg_memory_usage' in stats:
        metrics.append(("内存相关指标", None, True))  # 添加标题行
        metrics.extend([
            ('平均内存使用量(MB)', stats.get('avg_memory_usage', 0), False),
            ('峰值内存使用量(MB)', stats.get('max_memory_usage', 0), False),
        ])
    
    # 添加分类标题：FPS相关
    if 'avg_fps' in stats:
        metrics.append(("FPS相关指标", None, True))  # 添加标题行
        metrics.extend([
            ('平均FPS(帧)', stats.get('avg_fps', 0), False),
            ('峰值FPS(帧)', stats.get('max_fps', 0), False),
        ])
    
    # 添加分类标题：GPU相关
    if 'avg_gpu' in stats:
        metrics.append(("GPU相关指标", None, True))  # 添加标题行
        metrics.extend([
            ('平均GPU使用率(%)', stats.get('avg_gpu', 0), False),
            ('峰值GPU使用率(%)', stats.get('max_gpu', 0), False),
        ])
    
    # 添加分类标题：进程相关
    if 'avg_threads' in stats or 'avg_handles' in stats:
        metrics.append(("进程相关指标", None, True))  # 添加标题行
        if 'avg_threads' in stats:
            metrics.extend([
                ('平均线程数', stats.get('avg_threads', 0), False),
                ('峰值线程数', stats.get('max_threads', 0), False),
            ])
        if 'avg_handles' in stats:
            metrics.extend([
                ('平均句柄数', stats.get('avg_handles', 0), False),
                ('峰值句柄数', stats.get('max_handles', 0), False),
            ])
    
    # 添加分类标题：磁盘I/O相关
    if 'avg_disk_read_rate' in stats or 'avg_disk_write_rate' in stats:
        metrics.append(("磁盘I/O相关指标", None, True))  # 添加标题行
        if 'avg_disk_read_rate' in stats:
            metrics.extend([
                ('平均磁盘读取速率(MB/s)', stats.get('avg_disk_read_rate', 0), False),
                ('峰值磁盘读取速率(MB/s)', stats.get('max_disk_read_rate', 0), False),
            ])
        if 'avg_disk_write_rate' in stats:
            metrics.extend([
                ('平均磁盘写入速率(MB/s)', stats.get('avg_disk_write_rate', 0), False),
                ('峰值磁盘写入速率(MB/s)', stats.get('max_disk_write_rate', 0), False),
            ])
    
    # 添加分类标题：网络I/O相关
    if 'avg_net_sent_rate' in stats or 'avg_net_recv_rate' in stats:
        metrics.append(("网络I/O相关指标", None, True))  # 添加标题行
        if 'avg_net_sent_rate' in stats:
            metrics.extend([
                ('平均网络发送速率(MB/s)', stats.get('avg_net_sent_rate', 0), False),
                ('峰值网络发送速率(MB/s)', stats.get('max_net_sent_rate', 0), False),
            ])
        if 'avg_net_recv_rate' in stats:
            metrics.extend([
                ('平均网络接收速率(MB/s)', stats.get('avg_net_recv_rate', 0), False),
                ('峰值网络接收速率(MB/s)', stats.get('max_net_recv_rate', 0), False),
            ])
    
    # 添加分类标题：基础信息
    metrics.append(("基础信息", None, True))  # 添加标题行
    metrics.append(('数据点数量', data_length, False))
    
    # 写入统计数据并设置样式
    row_offset = stats_row + 1
    for label, value, is_title in metrics:
        if is_title:  # 如果是分类标题
            # 空一行，使布局更加清晰
            row_offset += 1
            title_cell = ws.cell(row=row_offset, column=stats_col, value=label)
            title_cell.font = openpyxl.styles.Font(bold=True)
            title_cell.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            # 合并标题单元格
            ws.merge_cells(start_row=row_offset, start_column=stats_col, end_row=row_offset, end_column=stats_col + 1)
        else:  # 如果是数据行
            label_cell = ws.cell(row=row_offset, column=stats_col, value=label)
            value_cell = ws.cell(row=row_offset, column=stats_col + 1, value=round(value, 2))
            
            # 设置标签和值单元格的样式
            label_cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            value_cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            
            # 交替行背景色
            if row_offset % 2 == 0:
                light_fill = openpyxl.styles.PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                label_cell.fill = light_fill
                value_cell.fill = light_fill
        
        row_offset += 1
    
    # 添加边框
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    for row in range(stats_row, row_offset):
        for col in range(stats_col, stats_col + 2):
            ws.cell(row=row, column=col).border = thin_border
    
    return row_offset

def add_charts_to_sheet(ws, data_rows_count, stats_last_row, has_gpu):
    """添加图表到工作表"""
    try:
        # 将图表起始位置设为统计信息下方2行
        chart_start_row = stats_last_row + 2
        
        # 创建CPU和内存使用图表
        performance_chart = create_performance_chart(ws, data_rows_count, has_gpu)
        ws.add_chart(performance_chart, f'A{chart_start_row}')
        
        # 各图表之间的间距调整为20行
        chart_spacing = 20
        
        # 创建FPS图表
        fps_chart = create_fps_chart(ws, data_rows_count)
        ws.add_chart(fps_chart, f'A{chart_start_row + chart_spacing}')
        
        # 创建线程和句柄图表
        threads_chart = create_threads_handles_chart(ws, data_rows_count)
        ws.add_chart(threads_chart, f'A{chart_start_row + chart_spacing * 2}')
        
        # 创建磁盘IO图表
        disk_io_chart = create_disk_io_chart(ws, data_rows_count)
        ws.add_chart(disk_io_chart, f'A{chart_start_row + chart_spacing * 3}')
        
        # 创建网络IO图表
        network_io_chart = create_network_io_chart(ws, data_rows_count)
        ws.add_chart(network_io_chart, f'A{chart_start_row + chart_spacing * 4}')
    except Exception as e:
        logger.error(f"添加图表时发生错误: {str(e)}")
        # 不中断执行，确保基本数据能正常保存

def create_performance_chart(ws, data_rows_count, has_gpu):
    """创建性能图表(CPU、内存、GPU)"""
    chart = LineChart()
    chart.title = '性能监控 (CPU/内存/GPU)'
    chart.y_axis.title = '使用率/使用量'
    chart.x_axis.title = '时间'
    
    # 仅当有数据行时才创建图表
    if data_rows_count <= 1:
        logger.warning("数据行数不足，无法创建性能图表")
        return chart
    
    try:
        # CPU使用率
        cpu_data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=data_rows_count)
        cpu_all_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=data_rows_count)
        memory_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=data_rows_count)
        
        chart.add_data(cpu_data, titles_from_data=True)
        chart.add_data(cpu_all_data, titles_from_data=True)
        chart.add_data(memory_data, titles_from_data=True)
        
        # GPU数据如果可用
        gpu_col = 6
        if has_gpu:
            gpu_data = Reference(ws, min_col=gpu_col, min_row=1, max_col=gpu_col, max_row=data_rows_count)
            chart.add_data(gpu_data, titles_from_data=True)
        
        # 时间轴
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count)
        chart.set_categories(cats)
        
        # 图表样式
        chart.style = 10
        chart.height = 15
        chart.width = 30
    except Exception as e:
        logger.error(f"创建性能图表时出错: {str(e)}")
    
    return chart

def create_fps_chart(ws, data_rows_count):
    """创建FPS图表"""
    chart = LineChart()
    chart.title = 'FPS监控'
    chart.y_axis.title = '帧率'
    chart.x_axis.title = '时间'
    
    # 仅当有数据行时才创建图表
    if data_rows_count <= 1:
        logger.warning("数据行数不足，无法创建FPS图表")
        return chart
    
    try:
        # FPS数据
        fps_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=data_rows_count)
        chart.add_data(fps_data, titles_from_data=True)
        
        # 时间轴
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count)
        chart.set_categories(cats)
        
        # 图表样式
        chart.style = 10
        chart.height = 15
        chart.width = 30
    except Exception as e:
        logger.error(f"创建FPS图表时出错: {str(e)}")
    
    return chart

def create_threads_handles_chart(ws, data_rows_count):
    """创建线程和句柄图表"""
    chart = LineChart()
    chart.title = '线程和句柄监控'
    chart.y_axis.title = '数量'
    chart.x_axis.title = '时间'
    
    # 仅当有数据行时才创建图表
    if data_rows_count <= 1:
        logger.warning("数据行数不足，无法创建线程和句柄图表")
        return chart
    
    try:
        # 获取线程和句柄列
        threads_col = 7  # 线程列
        handles_col = 8  # 句柄列
        
        # 添加线程和句柄数据
        threads_data = Reference(ws, min_col=threads_col, min_row=1, max_col=threads_col, max_row=data_rows_count)
        handles_data = Reference(ws, min_col=handles_col, min_row=1, max_col=handles_col, max_row=data_rows_count)
        
        chart.add_data(threads_data, titles_from_data=True)
        chart.add_data(handles_data, titles_from_data=True)
        
        # 时间轴
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count)
        chart.set_categories(cats)
        
        # 图表样式
        chart.style = 10
        chart.height = 15
        chart.width = 30
    except Exception as e:
        logger.error(f"创建线程和句柄图表时出错: {str(e)}")
    
    return chart

def create_disk_io_chart(ws, data_rows_count):
    """创建磁盘I/O图表"""
    chart = LineChart()
    chart.title = '磁盘I/O监控'
    chart.y_axis.title = '速率(MB/s)'
    chart.x_axis.title = '时间'
    
    # 仅当有数据行时才创建图表
    if data_rows_count <= 1:
        logger.warning("数据行数不足，无法创建磁盘I/O图表")
        return chart
    
    try:
        # 定位磁盘读取和写入速率列
        disk_read_col = 9  # 磁盘读取速率列
        disk_write_col = 10  # 磁盘写入速率列
        
        # 添加磁盘I/O数据
        disk_read_data = Reference(ws, min_col=disk_read_col, min_row=1, max_col=disk_read_col, max_row=data_rows_count)
        disk_write_data = Reference(ws, min_col=disk_write_col, min_row=1, max_col=disk_write_col, max_row=data_rows_count)
        
        chart.add_data(disk_read_data, titles_from_data=True)
        chart.add_data(disk_write_data, titles_from_data=True)
        
        # 时间轴
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count)
        chart.set_categories(cats)
        
        # 图表样式
        chart.style = 10
        chart.height = 15
        chart.width = 30
    except Exception as e:
        logger.error(f"创建磁盘I/O图表时出错: {str(e)}")
    
    return chart

def create_network_io_chart(ws, data_rows_count):
    """创建网络I/O图表"""
    chart = LineChart()
    chart.title = '网络I/O监控'
    chart.y_axis.title = '速率(MB/s)'
    chart.x_axis.title = '时间'
    
    # 仅当有数据行时才创建图表
    if data_rows_count <= 1:
        logger.warning("数据行数不足，无法创建网络I/O图表")
        return chart
    
    try:
        # 定位网络发送和接收速率列
        net_sent_col = 11  # 网络发送速率列
        net_recv_col = 12  # 网络接收速率列
        
        # 添加网络I/O数据
        net_sent_data = Reference(ws, min_col=net_sent_col, min_row=1, max_col=net_sent_col, max_row=data_rows_count)
        net_recv_data = Reference(ws, min_col=net_recv_col, min_row=1, max_col=net_recv_col, max_row=data_rows_count)
        
        chart.add_data(net_sent_data, titles_from_data=True)
        chart.add_data(net_recv_data, titles_from_data=True)
        
        # 时间轴
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows_count)
        chart.set_categories(cats)
        
        # 图表样式
        chart.style = 10
        chart.height = 15
        chart.width = 30
    except Exception as e:
        logger.error(f"创建网络I/O图表时出错: {str(e)}")
    
    return chart

def adjust_column_widths(ws):
    """调整工作表列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # 设置最小列宽，确保有足够的空间
        adjusted_width = max(12, (max_length + 2) * 1.2)
        # 限制最大列宽
        adjusted_width = min(adjusted_width, 50)
        ws.column_dimensions[column_letter].width = adjusted_width 